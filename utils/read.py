# Django
from django.db.models import Sum
# settings
from django.conf import settings
# openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
# records
from apps.records.models import Sucursales, UnidadesOperativas, Almacenes, Tiendas, Encargados, Supervisores
# users
from apps.users.models import Usuarios
# seeds
from apps.seeds.models import Pedidos, OrdenesSuministro, Movimientos, Productos
# services
from apps.services.models import Partidas, CEGAPS, Contratos, Facturas, CEGAPSPolizas, ExpedientePolizas
# movements
from apps.movements.models import MovimientosMercancias, ProductosMercancias
# zipfile
import zipfile
import os
import io

def stores_metadata():
    wb = load_workbook('stores_metadata.xlsx')
    ws = wb.active
    counter = 0
    for row in ws.iter_rows(min_row=2):
        counter += 1
        print(counter)
        clave_sucursal = str(row[0].value).strip()
        clave_unidad = str(row[2].value).strip()
        clave_almacen = str(row[5].value).strip()
        clave_tienda = str(row[6].value).strip()
        fecha = row[7].value
        encargado = row[8].value
        supervisor = row[9].value
        latitud = row[10].value
        longitud = row[11].value
        
        sucursal = Sucursales.objects.get(clave=clave_sucursal)
        unidad = UnidadesOperativas.objects.get(clave=clave_unidad, sucursal=sucursal)
        almacen = Almacenes.objects.get(clave=clave_almacen, unidad=unidad)
        stores = Tiendas.objects.filter(clave=clave_tienda, almacen=almacen)
        sup, created = Supervisores.objects.get_or_create(nombre=supervisor)
        enc, created = Encargados.objects.get_or_create(nombre=encargado)
        
        if stores.exists():
            for s in stores:        
                s.fecha_apertura = fecha
                s.supervisor = sup
                s.encargado = enc
                s.coords = {'latitud': latitud, 'longitud': longitud}
                s.save()
        else:
            s2 = Tiendas.objects.create(
                nombre=' ',
                clave=clave_tienda,
                fecha_apertura=fecha,
                coords={'latitud': latitud, 'longitud': longitud},
                importe=0.0,
                supervisor=sup,
                encargado=enc,
                almacen=almacen
            )
            print(s2.clave, s2.almacen.clave)

def write_pagares():
    wb = load_workbook('pagares_tiendas.xlsx')
    ws = wb.active
    zf = zipfile.ZipFile(os.path.join(settings.FILE_STRUCTURE_ROOT, 'PAGARES_OPERACION.zip'), 'w')
    output = io.BytesIO()
    counter = 1
    for row in ws.iter_rows(min_row=2):
        print(counter)
        nombre_unidad = str(row[2].value).strip()
        clave_almacen = str(row[5].value).strip()
        clave_tienda = str(row[9].value).strip()
        unidad = UnidadesOperativas.objects.get(nombre__exact=nombre_unidad)
        almacen = Almacenes.objects.get(clave=clave_almacen, unidad=unidad)
        tienda = almacen.tiendas.filter(clave=clave_tienda)
        
        if tienda.exists():
            tienda = tienda.first()
            exp = tienda.exp.first()
            if exp is not None:
                pagares = exp.pagares.all()
                pagares_22 = pagares.filter(anio='2022').last()
                pagares_23 = pagares.filter(anio='2023')
                pagare = pagares_23.last() or pagares_22                
                if pagare is not None:
                    name, extension = os.path.splitext(pagare.archivo.name)
                    path = os.path.join(tienda.almacen.unidad.nombre.strip(), tienda.almacen.nombre.strip(), tienda.clave.strip(), f'{name}{extension}')
                    try:
                        zf.write(pagare.archivo.path, path)
                        transfers = exp.transfers.all()
                        sum_trans = transfers.aggregate(costo=Sum('precio_monto'), venta=Sum('precio_venta'))
                        row[14].font = Font(name='Calibri', size=11, color='000000')
                        row[14].value = sum_trans['costo']
                        row[15].font = Font(name='Calibri', size=11, color='000000')
                        row[15].value = sum_trans['venta']
                        row[16].font = Font(name='Calibri', size=11, color='000000')
                        row[16].value = pagare.precio_venta
                        row[17].font = Font(name='Calibri', size=11, color='000000')
                        row[17].hyperlink = path
                        row[17].value = path
                        counter2 = 19
                        for t in transfers:
                            name, extension = os.path.splitext(t.archivo.name)
                            path = os.path.join(tienda.almacen.unidad.nombre.strip(), tienda.almacen.nombre.strip(), tienda.clave.strip(), f'{name}{extension}')
                            zf.write(t.archivo.path, path)
                            ws.cell(row=1, column=counter2).font = Font(name='Calibri', size=11, color='ffffffff')
                            ws.cell(row=1, column=counter2).fill = PatternFill(start_color='ff235b4e', end_color='ff235b4e', fill_type = "solid")
                            ws.cell(row=1, column=counter2).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=1, column=counter2).value = 'Archivo Transferencia'                            
                            ws.cell(row=counter+1, column=counter2).font = Font(name='Calibri', size=11, color='000000')
                            ws.cell(row=counter+1, column=counter2).hyperlink = path
                            ws.cell(row=counter+1, column=counter2).value = path
                            counter2 += 1
                    except Exception as e:
                        print(e, tienda.almacen.clave, tienda.clave)
        counter += 1
    wb.save(output)
    zf.writestr(f'PAGARES_TIENDAS.xlsx', output.getvalue())
    zf.close()
    
def read_cegaps2():
    wb = load_workbook('polizas_cegaps2.xlsx')
    ws = wb.active
    counter = 1
    for row in ws.iter_rows(min_row=2):
        nombre_unidad = str(row[12].value).strip()
        fecha_poliza = str(row[2].value).strip()
        cargo = row[9].value
        abono = row[10].value
        folio_poliza = row[11].value
        folio_cegap = str(row[15].value).strip()
        clave_partida = str(row[14].value).strip()
        partida_nombre = 'S/N'
        
        unidad = UnidadesOperativas.objects.get(nombre=nombre_unidad)
        partida, created = Partidas.objects.get_or_create(clave=clave_partida, defaults={'nombre': partida_nombre})
        cegap, created = CEGAPSPolizas.objects.get_or_create(
            folio_cegap=folio_cegap,
            partida=partida,
            unidad=unidad, 
            defaults={'prefolio': '', 'fecha': '', 'tipo': '', 'beneficiario': '', 'concepto': '', 'existe': True, 'periodo': 'p2'}
        )
        poliza = ExpedientePolizas.objects.create(fecha=fecha_poliza, folio_poliza=folio_poliza, cargo=cargo, abono=abono, cegap=cegap)
        print(counter)
        counter += 1

def read_cegaps():
    wb = load_workbook('polizas_cegaps.xlsx')
    ws = wb.active
    counter = 1
    for row in ws.iter_rows(min_row=2):
        clave_sucursal = str(row[0].value).strip()
        clave_unidad = str(row[1].value).strip()
        nombre_unidad = str(row[2].value).strip()
        prefolio = str(row[3].value).strip()
        folio_cegap = str(row[4].value).strip()
        existe = row[5].value
        fecha = row[6].value
        tipo = str(row[9].value).strip()
        beneficiario = str(row[11].value).strip()
        concepto = str(row[34].value).strip()
        clave_partida = str(row[20].value).strip()
        partida_nombre = 'S/N'
        
        sucursal = Sucursales.objects.get(clave=clave_sucursal)
        unidad, created = UnidadesOperativas.objects.get_or_create(clave=clave_unidad, sucursal=sucursal, defaults={'nombre':nombre_unidad})
        partida, created = Partidas.objects.get_or_create(clave=clave_partida, defaults={'nombre': partida_nombre})
        cegap, created = CEGAPSPolizas.objects.get_or_create(
            folio_cegap=folio_cegap,
            partida = partida,
            unidad = unidad, 
            defaults= {
                'prefolio': prefolio,
                'fecha': fecha,
                'tipo': tipo,
                'beneficiario': beneficiario,
                'concepto': concepto,
                'existe': True if existe == 'si' else False
            }
        )
        print(counter)
        counter += 1

def write_merca():
    wb = load_workbook('merca2.xlsx')
    ws = wb.active
    counter = 0
    print('hold')
    for row in ws.iter_rows(min_row=1):
        counter += 1
        print(counter)
        clave_sucursal = str(row[0].value).strip()
        nombre_sucursal = str(row[1].value).strip()
        clave_uo = str(row[2].value).strip()
        nombre_uo = str(row[3].value).strip()
        clave_almacen = str(row[4].value).strip()
        nombre_almacen = str(row[5].value).strip()
        clave_movimiento = str(row[6].value).strip()
        canal = str(row[7].value).strip()
        orden = str(row[8].value).strip()
        folio_movimiento = str(row[9].value).strip()
        fecha_movimiento = row[10].value
        clave_familia = str(row[11].value).strip()
        nombre_familia = str(row[12].value).strip()
        clave_producto = str(row[13].value).strip()
        nombre_producto = str(row[14].value).strip()
        clave_articulo = str(row[15].value).strip()
        nombre_articulo = str(row[16].value).strip()
        unidad_medida = str(row[17].value).strip()
        presentacion = str(row[18].value).strip()
        volumen = row[19].value
        importe = row[20].value
        folio_fiscal = row[21].value
        
        try:
            almacen = Almacenes.objects.get(clave=clave_almacen)
        except Almacenes.DoesNotExist:
            sucursal = Sucursales.objects.get(clave=clave_sucursal)
            uo = UnidadesOperativas.objects.get(clave=clave_uo, sucursal=sucursal)
            almacen = Almacenes.objects.create(unidad=uo, nombre=nombre_almacen, clave=clave_almacen)
            create_users(almacen)
            print(clave_almacen, 'creado')
            
        obj3, created = ProductosMercancias.objects.get_or_create(
            clave_articulo=clave_articulo, 
            defaults={
                'clave': clave_producto, 
                'nombre': nombre_producto, 
                'clave_familia': clave_familia, 
                'familia': nombre_familia, 
                'nombre_articulo': nombre_articulo, 
                'unidad_medida': unidad_medida, 
                'presentacion': presentacion
            }
        )
        
        obj4 = MovimientosMercancias.objects.create(
            folio_doc=folio_movimiento, 
            almacen=almacen, 
            unidad=almacen.unidad,
            producto=obj3, 
            fecha=fecha_movimiento, 
            volumen=volumen, 
            importe=importe, 
            clave=clave_movimiento,
            orden=orden,
            canal=canal,
            folio_fiscal=folio_fiscal
        )
    

def write_stores():
    wb = load_workbook('stores.xlsx')
    ws = wb.active
    zf = zipfile.ZipFile(os.path.join(settings.FILE_STRUCTURE_ROOT, 'PAGARES_OPERACION.zip'), 'w')
    output = io.BytesIO()
    counter = 0
    for row in ws.iter_rows(min_row=2):
        counter += 1
        print(counter)
        clave_region = str(row[2].value).strip()
        clave_unidad = str(row[4].value).strip()
        clave_almacen = str(row[8].value).strip()
        clave_tienda = str(row[17].value).strip()
        store = Tiendas.objects.filter(clave=clave_tienda, almacen__clave=clave_almacen, almacen__unidad__clave=clave_unidad, almacen__unidad__sucursal__clave=clave_region)
        if store.exists():
            row[19].value = 'Sí'
            store = store.first()
            exp = store.exp.first()
            if exp is not None:
                pagares = exp.pagares.all()
                pagares_20 = pagares.filter(anio='2020').last()
                pagares_21 = pagares.filter(anio='2021').last()
                pagares_22 = pagares.filter(anio='2022').last()
                pagares_23 = pagares.filter(anio='2023')
                if row[1].value == 'Cerrada':
                    pagare = pagares_23.last() or pagares_22 or pagares_21 or pagares_20
                else:
                    pagare = pagares_23.last() or pagares_22                    
                row[20].value = pagares.count()
                row[21].value = pagares_23.count()
                row[22].value = exp.get_operacion_display() if exp.operacion is not None else 'No especificado'
                if pagare is not None:
                    name, extension = os.path.splitext(pagare.archivo.name)
                    path = os.path.join(store.almacen.unidad.nombre.strip(), store.almacen.nombre.strip(), store.clave.strip(), f'{name}{extension}')
                    try:
                        # zf.write(pagare.archivo.path, path)
                        row[23].hyperlink = path
                        row[23].value = path
                    except Exception as e:
                        row[23].value = ''
                        print(e, store.almacen.clave, store.clave)
                    row[24].value = pagare.anio
                    row[25].value = pagare.precio_venta
        else:
            row[19].value = 'No'
    wb.save(output)
    zf.writestr(f'PAGARES.xlsx', output.getvalue())
    zf.close()

def read_services():
    wb = load_workbook('services.xlsx')
    ws = wb.active
    counter = 1
    for row in ws.iter_rows(min_row=2):
        clave_sucursal = str(row[0].value).strip()
        clave_unidad = str(row[1].value).strip()
        prefolio = str(row[3].value).strip()
        folio_cegap = str(row[4].value).strip()
        fecha = row[5].value
        tipo = str(row[6].value).strip()
        beneficiario = str(row[7].value).strip()
        concepto = str(row[8].value).strip()
        tipo_pago = str(row[9].value).strip()
        cheque = str(row[10].value).strip()
        fecha_pago = row[11].value
        clave_partida = str(row[12].value).strip()
        partida_nombre = str(row[13].value).strip()
        folio_cecopre = str(row[14].value).strip()
        importe = row[15].value
        importe_iva = row[16].value
        importe_total = row[17].value
        folio_fiscal = str(row[18].value).lower()
        rfc = str(row[19].value).strip()
        folio_contrato = str(row[20].value).strip()
        
        sucursal = Sucursales.objects.get(clave=clave_sucursal)
        unidad = UnidadesOperativas.objects.get(clave=clave_unidad, sucursal=sucursal)
        partida, created = Partidas.objects.get_or_create(clave=clave_partida, defaults={'nombre': partida_nombre})
        cegap, created = CEGAPS.objects.get_or_create(
            folio_cegap=folio_cegap,
            partida = partida,
            unidad = unidad, 
            defaults= {
                'prefolio': prefolio,
                'fecha': fecha,
                'tipo': tipo,
                'beneficiario': beneficiario,
                'concepto': concepto,
                'tipo_pago': tipo_pago,
                'cheque': cheque,
                'fecha_pago': fecha_pago,
                'folio_cecopre': folio_cecopre
            }
        )
        if not folio_contrato == 'None':
            print(folio_contrato)
            Contratos.objects.get_or_create(clave=folio_contrato, cegap=cegap)
        Facturas.objects.create(
            importe=importe,
            importe_iva=importe_iva,
            importe_total=importe_total,
            rfc_proveedor=rfc,
            folio_fiscal=folio_fiscal,
            cegap=cegap
        )
        print(counter)
        counter += 1

def read_seed():
    wb = load_workbook('seeds2.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        clave_sucursal = str(row[0].value).strip()
        clave_uo = str(row[2].value).strip()
        nombre_almacen = str(row[5].value).strip()
        clave_almacen = str(row[4].value).strip()
        folio_pedido = str(row[10].value).strip()
        folio_orden = str(row[12].value).strip()
        clave_producto = str(row[21].value).strip()
        nombre_producto = str(row[22].value).strip()
        clave_familia = str(row[19].value).strip()
        nombre_familia = str(row[20].value).strip()
        clave_articulo = str(row[23].value).strip()
        nombre_articulo = str(row[24].value).strip()
        unidad_medida = str(row[27].value).strip()
        presentacion = str(row[28].value).strip()
        folio_movimiento = str(row[15].value).strip()
        fecha_movimiento = row[17].value
        volumen_movimiento = row[29].value
        importe_movimiento = row[31].value
        clave_movimiento = str(row[6].value).strip()
        
        print(clave_almacen)
        try:
            almacen = Almacenes.objects.get(clave=clave_almacen)
        except Almacenes.DoesNotExist:
            sucursal = Sucursales.objects.get(clave=clave_sucursal)
            uo = UnidadesOperativas.objects.get(clave=clave_uo, sucursal=sucursal)
            almacen = Almacenes.objects.create(unidad=uo, nombre=nombre_almacen, clave=clave_almacen)
            create_users(almacen)
            print(clave_almacen, 'creado')
            
        obj1, created = Pedidos.objects.get_or_create(folio_doc=folio_pedido, almacen=almacen)
        obj2, created = OrdenesSuministro.objects.get_or_create(folio_doc=folio_orden, pedido=obj1)
        obj3, created = Productos.objects.get_or_create(clave_articulo=clave_articulo, defaults={'clave': clave_producto, 'nombre': nombre_producto, 'clave_familia': clave_familia, 'familia': nombre_familia, 'nombre_articulo': nombre_articulo, 'unidad_medida': unidad_medida, 'presentacion': presentacion})
        obj4 = Movimientos.objects.create(folio_doc=folio_movimiento, orden=obj2, almacen=almacen, producto=obj3, fecha=fecha_movimiento, volumen=volumen_movimiento, importe=importe_movimiento, clave=clave_movimiento)
    

def read_xlsx():
    wb = load_workbook('tiendas.xlsx')
    ws = wb.active
    objs = []
    for row in ws.iter_rows(min_row=2):
        suc, created = Sucursales.objects.get_or_create(clave=row[0].value, defaults={'nombre': row[1].value})
        uop, created = UnidadesOperativas.objects.get_or_create(clave=row[2].value, nombre=row[3].value, defaults={'sucursal': suc})
        alm, created = Almacenes.objects.get_or_create(clave=row[4].value, nombre=row[5].value, defaults={'unidad': uop})
        Tiendas.objects.create(clave=row[8].value, nombre=row[9].value, almacen=alm, importe=row[11].value)

def create_users(alm):
    user = Usuarios.objects.create(
        username=alm.clave,
        email='{}@expediente.com'.format(alm.clave),
        phone_number = '1234567890',
        full_name = 'Usuario de almacén'
    )
    user.set_password('Temporal123')
    user.save()
    alm.usuario = user
    alm.save()
    # for alm in Almacenes.objects.all():
    #     user = Usuarios.objects.create(
    #         username=alm.clave,
    #         email='{}@expediente.com'.format(alm.clave),
    #         phone_number = '1234567890',
    #         full_name = 'Usuario de almacén'
    #     )
    #     user.set_password('Temporal123')
    #     user.save()
    #     alm.usuario = user
    #     alm.save()

def create_users_uo():
    uos = UnidadesOperativas.objects.all()
    for uo in uos:    
        user = Usuarios.objects.create(
            username= f'uo{uo.sucursal.clave.strip()}_{uo.clave.strip()}',
            email=f'uo{uo.sucursal.clave.strip()}_{uo.clave.strip()}@expediente.com',
            phone_number = '1234567890',
            full_name = 'Usuario de Unidad Operativa'
        )
        user.set_password('Temporal123')
        user.user_type = 'unidad'
        user.save()
        uo.usuario = user
        uo.save()