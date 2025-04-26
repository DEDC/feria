# Django
from django.http import HttpResponse
from django.db.models import Q, Sum, Case, When, Value, CharField, Count
import zipfile
import os
# places
from apps.places.models import Lugares, Pagos, ProductosExtras, Solicitudes
# openpyxl
from openpyxl import load_workbook
import pytz
# utils
from utils.tools import get_difference_by_different_keys, flatten_nested_dicts

# Especificar la zona horaria deseada
zona_horaria = pytz.timezone("America/Mexico_City")


def get_docs():
    zf = zipfile.ZipFile(os.path.join('', 'SISCOMFERIA_DOCUMENTACION.zip'), 'w')
    counter = 1
    for sl in Solicitudes.objects.filter(solicitud_pagos__pagado=True):
        print(counter)
        if sl.identificacion:
            try:
                zf.write(sl.identificacion.path, f'{sl.folio}_{sl.nombre}/{os.path.basename(sl.identificacion.name)}')
            except Exception as e:
                pass
                print(e)
        
        if sl.acta_constitutiva:
            try:
                zf.write(sl.acta_constitutiva.path, f'{sl.folio}_{sl.nombre}/{os.path.basename(sl.acta_constitutiva.name)}')
            except Exception as e:
                pass
                print(e)
        
        if sl.comprobante_domicilio:
            try:
                zf.write(sl.comprobante_domicilio.path, f'{sl.folio}_{sl.nombre}/{os.path.basename(sl.comprobante_domicilio.name)}')
            except Exception as e:
                pass
                print(e)
        
        if sl.constancia_fiscal:
            try:
                zf.write(sl.constancia_fiscal.path, f'{sl.folio}_{sl.nombre}/{os.path.basename(sl.constancia_fiscal.name)}')
            except Exception as e:
                pass
                print(e)
        
        if sl.curp:
            try:
                zf.write(sl.curp.path, f'{sl.folio}_{sl.nombre}/{os.path.basename(sl.curp.name)}')
            except Exception as e:
                pass
                print(e)
        counter+=1
    zf.close()

def get_stands_report(places_dict):
    wb = load_workbook('static/docs/reporte_locales.xlsx')
    ws = wb.get_sheet_by_name('LOCALES')
    counter = 3
    zone_choices = Lugares.zona.field.choices
    # get the queryset values using when case
    selected_places = Lugares.objects.all().select_related('solicitud_lugar').prefetch_related('extras').annotate(
        zona_display=Case(*[When(zona=choice[0], then=Value(choice[1])) for choice in zone_choices], output_field=CharField(), default=Value('Desconocido')),
        is_paid=Case(When(Q(tpay_pagado=True) | Q(caja_pago=True) | Q(transfer_pago=True), then=Value('Sí')), output_field=CharField(), default=Value('No')),
        permisos=Count('extras__tipo', filter=Q(extras__tipo='licencia_alcohol')),
        has_alcohol=Case(When(permisos__gt=1, then=Value("Sí")), default=Value("No"), output_field=CharField())
    ).values('folio', 'uuid', 'uuid_place', 'nombre', 'zona', 'zona_display', 'm2', 'precio', 'tramite_id', 'is_paid', 'has_alcohol', 'solicitud__folio', 'solicitud__uuid', 'solicitud__nombre')

    # convert queryset to mutable list
    selected_places = list(selected_places)
    metaplaces = places_dict.copy()
    for p in selected_places:
        if p['zona'] in metaplaces.keys():
            result = next((place for place in metaplaces[p['zona']]['places'] if place['uuid'] == str(p['uuid_place'])), None)
            if result:
                p['price'] = result['price']
                p['uuid_internal'] = result['uuid']
                p['concept_internal'] = result['concept']
                p['added_by'] = 'Mapa'
            else:
                p['added_by'] = 'Manual'
        else:
            p['added_by'] = 'Manual'
    
    # the list of dicts converts in one single list o dicts 
    base_places = flatten_nested_dicts(places_dict)                
    
    # get the difference beetwen base place to selected places on bd
    not_selected_places = get_difference_by_different_keys(base_places, selected_places, 'uuid', 'uuid_place')
    all_places = sorted(selected_places + not_selected_places, key=lambda x: x['zona_display'])

    for p in all_places:
        ws.cell(row=counter, column=1, value=p.get('nombre', p.get('text', 'No especificado')))
        ws.cell(row=counter, column=2, value=p.get('zona_display', 'No especificado'))
        ws.cell(row=counter, column=3, value=p.get('folio', 'Sin asignar'))
        ws.cell(row=counter, column=4, value=p.get('solicitud__folio', 'Sin asignar'))
        ws.cell(row=counter, column=5, value=p.get('solicitud__nombre', 'Sin asignar'))

        ws.cell(row=counter, column=9, value=p.get('has_alcohol', 'Sin asignar'))
        ws.cell(row=counter, column=10, value=p.get('is_paid', 'Sin asignar'))

        # if hasattr(p.solicitud, 'comercio'):
        #     c = p.solicitud.comercio
        #     ws.cell(row=counter, column=7, value=c.nombre)
        #     ws.cell(row=counter, column=8, value=c.get_giro_display())
        # else:
        #     ws.cell(row=counter, column=7, value='No especificado')
        #     ws.cell(row=counter, column=8, value='No especificado')

        counter+=1
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    wb.save(response)
    response['Content-Disposition'] = 'attachment; filename=REPORTE_LOCALES_FERIA_2025.xlsx'
    return response

def get_requests_report():
    wb = load_workbook('static/docs/reporte_solicitudes.xlsx')
    ws = wb.get_sheet_by_name('SOLICITUDES')
    requests = Solicitudes.objects.all().order_by('fecha_reg')
    counter = 3
    for r in requests:
        # tiendas info
        ws.cell(row=counter, column=1, value=r.folio)
        ws.cell(row=counter, column=2, value=r.get_estatus_display() if r.estatus else 'Sin asignar')
        ws.cell(row=counter, column=3, value=r.fecha_reg.strftime('%d-%m-%Y'))
        ws.cell(row=counter, column=4, value=r.usuario.get_full_name())
        ws.cell(row=counter, column=5, value=str(r.usuario.phone_number))
        ws.cell(row=counter, column=6, value= 'Sí' if r.factura else 'No')
        ws.cell(row=counter, column=7, value= 'Sí' if r.regimen_fiscal else 'No')
        ws.cell(row=counter, column=8, value=r.nombre)
        ws.cell(row=counter, column=9, value=r.nombre_replegal if r.nombre_replegal is not None else '')
        ws.cell(row=counter, column=10, value=r.rfc_txt.upper() if r.rfc_txt is not None else '')
        ws.cell(row=counter, column=11, value=r.curp_txt.upper() if r.curp_txt is not None else '')
        ws.cell(row=counter, column=12, value=int(r.cantidad_espacios))
        ws.cell(row=counter, column=13, value= 'Sí' if r.mas_espacios else 'No')
        ws.cell(row=counter, column=14, value=r.calle)
        ws.cell(row=counter, column=15, value=r.no_calle)
        ws.cell(row=counter, column=16, value=r.colonia)
        ws.cell(row=counter, column=17, value=r.codigo_postal)
        ws.cell(row=counter, column=18, value=r.get_estado_display())
        ws.cell(row=counter, column=19, value=r.municipio)
        if hasattr(r, 'comercio'):
            c = r.comercio
            ws.cell(row=counter, column=20, value=c.folio)
            ws.cell(row=counter, column=21, value=c.get_estatus_display() if c.estatus else 'Sin asignar')
            ws.cell(row=counter, column=22, value=c.nombre)
            ws.cell(row=counter, column=23, value=c.get_giro_display())
            ws.cell(row=counter, column=24, value=c.descripcion)
            ws.cell(row=counter, column=25, value='Sí' if c.vende_alcohol else 'No')
        counter+=1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    wb.save(response)
    response['Content-Disposition'] = 'attachment; filename=REPORTE_SOLICITUDES_FERIA_2025.xlsx'
    return response


def get_report():
    wb = load_workbook('static/docs/layout_feria.xlsx')
    ws = wb.get_sheet_by_name('LOCALES')
    places = Lugares.objects.filter(Q(tpay_pagado=True) | Q(caja_pago=True) | Q(transfer_pago=True)).order_by('solicitud__folio', 'zona')
    # places = Lugares.objects.all().order_by('solicitud__folio', 'zona')
    counter = 4
    for p in places:
        ws.cell(row=counter, column=1).value = p.folio
        ws.cell(row=counter, column=2).value = p.nombre
        ws.cell(row=counter, column=3).value = p.get_zona_display()
        ws.cell(row=counter, column=4).value = p.fecha_reg.astimezone(zona_horaria).strftime("%Y-%m-%d %H:%M:%S") or 'No definido'
        ws.cell(row=counter, column=5).value = p.precio
        ws.cell(row=counter, column=6).value = p.m2
        ws.cell(row=counter, column=7).value = p.solicitud.nombre
        ws.cell(row=counter, column=8).value = p.solicitud.folio
        ws.cell(row=counter, column=9).value = p.solicitud.get_estatus_display()
        ws.cell(row=counter, column=10).value = p.solicitud.comercio.nombre if hasattr(p.solicitud, 'comercio') else 'Sin Comercio'
        ws.cell(row=counter, column=11).value = p.solicitud.comercio.get_giro_display() if hasattr(p.solicitud, 'comercio') else 'Sin Comercio'
        ws.cell(row=counter, column=12).value = p.solicitud.usuario.email
        ws.cell(row=counter, column=13).value = p.solicitud.usuario.phone_number
        ws.cell(row=counter, column=14).value = p.solicitud.usuario.get_full_name()
        payment = p.solicitud.solicitud_pagos.first()
        if p.tpay_pagado:
            status = "TPAY"
            if payment:
                status = payment.get_tipo_display()
            ws.cell(row=counter, column=15).value = 'Pagado' if p.tpay_pagado else 'No pagado'
            ws.cell(row=counter, column=16).value = status
            ws.cell(row=counter, column=17).value = p.fecha_mod.astimezone(zona_horaria).strftime("%Y-%m-%d") or 'No definido'
            ws.cell(row=counter, column=18).value = p.fecha_mod.astimezone(zona_horaria).strftime("%H:%M:%S") or 'No definido'
            ws.cell(row=counter, column=19).value = payment.validador if payment else 'No definido'
        if p.caja_pago:
            status = "TPAY"
            if payment:
                status = payment.get_tipo_display()
            ws.cell(row=counter, column=15).value = 'Pagado'
            ws.cell(row=counter, column=16).value = "CAJA"
            ws.cell(row=counter, column=17).value = p.fecha_mod.astimezone(zona_horaria).strftime("%Y-%m-%d") or 'No definido'
            ws.cell(row=counter, column=18).value = p.fecha_mod.astimezone(zona_horaria).strftime("%H:%M:%S") or 'No definido'
            ws.cell(row=counter, column=19).value = payment.validador if payment else 'No definido'
        if p.transfer_pago:
            ws.cell(row=counter, column=15).value = 'Pagado'
            ws.cell(row=counter, column=16).value = "Transferencia"
            ws.cell(row=counter, column=17).value = p.fecha_mod.astimezone(zona_horaria).strftime("%Y-%m-%d") or 'No definido'
            ws.cell(row=counter, column=18).value = p.fecha_mod.astimezone(zona_horaria).strftime("%H:%M:%S") or 'No definido'
            ws.cell(row=counter, column=19).value = payment.validador if payment else 'No definido'
        counter_column = 20
        for px in p.extras.all():
            ws.cell(row=counter, column=counter_column).value = px.get_tipo_display()
            counter_column += 1
            ws.cell(row=counter, column=counter_column).value = px.precio
            counter_column += 1
            ws.cell(row=counter, column=counter_column).value = 'Aplica para {}'.format(px.to_places)
            counter_column += 1
        counter += 1


    ws = wb.get_sheet_by_name('MONTOS')
    # n1
    ws['E4'] = Lugares.objects.filter(zona='n_1').count()
    ws['E8'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='n_1')))['total']
    ws['H4'] = Lugares.objects.filter(zona='n_1').aggregate(total=Sum('precio'))['total']
    ws['H8'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='n_1')))['total']
    # n3
    ws['O4'] = Lugares.objects.filter(zona='n_3').count()
    ws['O8'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='n_3')))['total']
    ws['R4'] = Lugares.objects.filter(zona='n_3').aggregate(total=Sum('precio'))['total']
    ws['R8'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='n_3')))['total']
    # z_a
    ws['E13'] = Lugares.objects.filter(zona='z_a').count()
    ws['E17'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='z_a')))['total']
    ws['H13'] = Lugares.objects.filter(zona='z_a').aggregate(total=Sum('precio'))['total']
    ws['H17'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='z_a')))['total']
    # z_b
    ws['O13'] = Lugares.objects.filter(zona='z_b').count()
    ws['O17'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='z_b')))['total']
    ws['R13'] = Lugares.objects.filter(zona='z_b').aggregate(total=Sum('precio'))['total']
    ws['R17'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='z_b')))['total']
    # z_c
    ws['E22'] = Lugares.objects.filter(zona='z_c').count()
    ws['E26'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='z_c')))['total']
    ws['H22'] = Lugares.objects.filter(zona='z_c').aggregate(total=Sum('precio'))['total']
    ws['H26'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='z_c')))['total']
    # z_d
    ws['O22'] = Lugares.objects.filter(zona='z_d').count()
    ws['O26'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar', filter=Q(solicitud__solicitud_lugar__zona='z_d')))['total']
    ws['R22'] = Lugares.objects.filter(zona='z_d').aggregate(total=Sum('precio'))['total']
    ws['R26'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio', filter=Q(solicitud__solicitud_lugar__zona='z_d')))['total']
    # terraza
    ws['E37'] = ProductosExtras.objects.filter(tipo='terraza').count()
    ws['E41'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar__extras', filter=Q(solicitud__solicitud_lugar__extras__tipo='terraza')))['total']
    ws['H37'] = ProductosExtras.objects.filter(tipo='terraza').aggregate(total=Sum('precio'))['total']
    ws['H41'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__extras__precio', filter=Q(solicitud__solicitud_lugar__extras__tipo='terraza')))['total']
    # terraza_grande
    ws['O37'] = ProductosExtras.objects.filter(tipo='terraza_grande').count()
    ws['O41'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar__extras', filter=Q(solicitud__solicitud_lugar__extras__tipo='terraza_grande')))['total']
    ws['R37'] = ProductosExtras.objects.filter(tipo='terraza_grande').aggregate(total=Sum('precio'))['total']
    ws['R41'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__extras__precio', filter=Q(solicitud__solicitud_lugar__extras__tipo='terraza_grande')))['total']
    # licencias_alcohol
    ws['E46'] = ProductosExtras.objects.filter(tipo='licencia_alcohol').count()
    ws['E50'] = Pagos.objects.filter(pagado=True).aggregate(total=Count('solicitud__solicitud_lugar__extras', filter=Q(solicitud__solicitud_lugar__extras__tipo='licencia_alcohol')))['total']
    ws['H46'] = ProductosExtras.objects.filter(tipo='licencia_alcohol').aggregate(total=Sum('precio'))['total']
    ws['H50'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__extras__precio', filter=Q(solicitud__solicitud_lugar__extras__tipo='licencia_alcohol')))['total']

    ws['Q29'] = Lugares.objects.all().aggregate(total=Sum('precio'))['total']
    ws['Q32'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__precio'))['total']

    ws['Q53'] = ProductosExtras.objects.all().aggregate(total=Sum('precio'))['total']
    ws['Q56'] = Pagos.objects.filter(pagado=True).aggregate(total=Sum('solicitud__solicitud_lugar__extras__precio'))['total']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    wb.save(response)
    response['Content-Disposition'] = 'attachment; filename=REPORTE_FERIA_2025.xlsx'
    return response