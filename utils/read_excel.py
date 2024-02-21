# Django
from django.db import IntegrityError
# openpyxl
from openpyxl import load_workbook
# users
from apps.users.models import Usuarios
# dates
from apps.dates.models import CitasAgendadas

def read_ambulantes():
    wb = load_workbook('static/docs/ambulantes.xlsx')
    ws = wb.active
    counter = 0
    
    for row in ws.iter_rows(min_row=2):
        counter += 1
        first_name = str(row[0].value).strip()
        last_name = str(row[1].value).strip()
        phone_number = str(row[2].value).strip()
        email = str(row[3].value).strip().lower()
        password = str(row[4].value).strip()
        date = str(row[5].value).strip()
        hour = str(row[6].value).strip()
        print(counter, first_name)
        try:
            usu = Usuarios.objects.create(email=email, first_name=first_name, last_name=last_name, phone_number=phone_number)
            usu.set_password(password)
            usu.save()
            date = CitasAgendadas.objects.get_or_create(fecha=date, hora=hour, usuario=usu)
        except IntegrityError:
            print(email, 'repetido')
            pass